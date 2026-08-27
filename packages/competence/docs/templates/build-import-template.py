"""Bring the HR employee-import template in line with the CSV column contract.

Run manually, from the repository root:

    python packages/competence/docs/templates/build-import-template.py

This is NOT part of the Node build or of CI. It exists because the template was originally hand-built, so when
CA-109 added two columns there was no way to regenerate it except by hand again.

It EXTENDS the existing workbook rather than rebuilding it. What goes stale when the contract changes is the column
set, the dropdowns and the cell formats; the Instructions prose is authored content, and re-expressing it as Python
string literals would make it harder to edit and would silently drop any formatting nobody thought to re-encode.

Requires openpyxl (a local dev tool only -- it is deliberately not a dependency of the competence package, whose
runtime dependencies are limited to core, web-framework and graphology).
"""

import pathlib
import openpyxl
from openpyxl.worksheet.datavalidation import DataValidation

TEMPLATE = pathlib.Path(__file__).with_name("employee-import-template.xlsx")
LAST_ROW = 401

# Columns appended by CA-109, in CSV-contract order. `work_site` is deliberately FREE TEXT: the valid codes are a
# per-deployment configuration document edited in Administration > Work Sites, so a dropdown baked into a file
# committed to a public repository would carry the demo codes and be wrong in every real install.
NEW_COLUMNS = ["work_site", "position_name"]

# `gender` is constrained to M/F on every write path as of CA-109, so the sheet should stop accepting free text.
GENDER_CHOICES = '"M,F"'

# Every existing dropdown on the Employees sheet shares this exact error/errorTitle wording, plus a promptTitle
# and prompt hint -- matched here so the new gender dropdown reads as one of the family, not a bolted-on extra.
VALIDATION_ERROR = "Pick a value from the list. The importer rejects anything else - it will not guess."
VALIDATION_ERROR_TITLE = "Not a permitted value"

# Where the Instructions sheet repeats the Employees header as a worked example (rows 30-31). Confirmed against
# the file rather than assumed: columns A/B there are a margin and a label column, not part of the table, so the
# example starts at C -- unlike the Employees sheet itself, where the same header starts at A.
EXAMPLE_HEADER_ROW = 30
EXAMPLE_FIRST_COLUMN = 3  # column C


def header_map(sheet):
    """Column name -> 1-based index, from row 1."""
    return {cell.value: cell.column for cell in sheet[1] if cell.value}


def append_missing_columns(sheet):
    columns = header_map(sheet)
    for name in NEW_COLUMNS:
        if name in columns:
            print(f"  {name}: already present, left alone")
            continue
        index = sheet.max_column + 1
        sheet.cell(row=1, column=index, value=name)
        # Copy the header style from an existing OPTIONAL column so the new ones read as optional (grey), not
        # required (dark blue). employment_status is the first optional column.
        source = sheet.cell(row=1, column=columns["employment_status"])
        target = sheet.cell(row=1, column=index)
        target.font = source.font.copy()
        target.fill = source.fill.copy()
        target.border = source.border.copy()
        target.alignment = source.alignment.copy()
        sheet.column_dimensions[target.column_letter].width = 22
        print(f"  {name}: added as column {target.column_letter}")


def constrain_gender(sheet):
    columns = header_map(sheet)
    letter = sheet.cell(row=1, column=columns["gender"]).column_letter
    target = f"{letter}2:{letter}{LAST_ROW}"
    for existing in list(sheet.data_validations.dataValidation):
        if str(existing.sqref) == target:
            print("  gender: validation already present, left alone")
            return
    validation = DataValidation(type="list", formula1=GENDER_CHOICES, allow_blank=True, showDropDown=False)
    validation.error = VALIDATION_ERROR
    validation.errorTitle = VALIDATION_ERROR_TITLE
    validation.promptTitle = "gender"
    validation.prompt = "Permitted: M, F"
    sheet.add_data_validation(validation)
    validation.add(target)
    print(f"  gender: dropdown added on {target}")


def refresh_instructions(book):
    """Extend the header/example pair on the Instructions sheet to cover the new columns.

    Only the header row (30) is touched -- the example data row (31) is authored content, and the existing
    example already leaves an optional column (gender) blank on purpose, which reads as deliberate ("blank means
    leave unchanged") rather than an oversight. Newly-appended header cells get the same grey "optional" look as
    their neighbours instead of landing as unstyled plain text.
    """
    sheet = book["Instructions"]
    employees = book["Employees"]
    columns = header_map(employees)
    names = [cell.value for cell in employees[1] if cell.value]
    optional_from = columns["employment_status"]  # 1-based column index on the Employees sheet
    style_source = sheet.cell(row=EXAMPLE_HEADER_ROW, column=EXAMPLE_FIRST_COLUMN + optional_from - 1)
    for offset, name in enumerate(names):
        cell = sheet.cell(row=EXAMPLE_HEADER_ROW, column=EXAMPLE_FIRST_COLUMN + offset, value=name)
        if offset + 1 >= optional_from:  # the optional zone, including any newly-appended column
            cell.font = style_source.font.copy()
            cell.fill = style_source.fill.copy()
            cell.border = style_source.border.copy()
            cell.alignment = style_source.alignment.copy()
    print(f"  Instructions: example header rewritten with {len(names)} columns")


def main():
    book = openpyxl.load_workbook(TEMPLATE)
    print(f"Updating {TEMPLATE.name}")
    append_missing_columns(book["Employees"])
    constrain_gender(book["Employees"])
    refresh_instructions(book)
    book.save(TEMPLATE)
    print("Saved. Open it in Excel or LibreOffice before committing -- see Task 10's Step 5 checklist.")


if __name__ == "__main__":
    main()
